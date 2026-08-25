"""Recruiting portal: bulk-import candidates from a LinkedIn Recruiter "Job
Applicant Report" .xlsx export. These are real applicants (source='linkedin'),
distinct from the public form (source='form') -- see Candidate.source.

Import never auto-creates a Match to a Requisition/project: fuzzy-matching
the export's Job Title/Hiring Project Title against however a recruiter
phrased an existing Requisition.title is unreliable. Imported candidates
land in the general pool and get added to a project manually via the
existing search "Add" control, same as any other candidate.
"""
import re

import openpyxl
from django.db import IntegrityError
from django.shortcuts import render

from .decorators import recruiter_required
from .models import Candidate, STATE_PROVINCE_CHOICES

# The export's first worksheet is empty in practice -- locate the data sheet
# by its header row instead of assuming a fixed sheet index/name.
EXPECTED_HEADER = 'First Name'

TITLE_KEYWORDS = [
    ('manager', 'Sales Manager'),
    ('sales development', 'SDR'),
    ('sdr', 'SDR'),
    ('business development', 'BDR'),
    ('bdr', 'BDR'),
    ('account executive', 'AE'),
]

# Best-effort match against a "How many years... : 12" style screening
# question -- this exact phrasing depends on whatever questions were
# configured on the job posting, so it's not guaranteed on every export.
YEARS_RE = re.compile(r'years[^:]*:\s*(\d+)', re.IGNORECASE)


def clean(value):
    if value is None:
        return ''
    text = str(value).strip()
    return '' if text.upper() == 'N/A' else text


def map_title(raw_title):
    low = raw_title.lower()
    for keyword, choice in TITLE_KEYWORDS:
        if keyword in low:
            return choice
    return 'Other'


def find_state_province(location_text):
    for _group, options in STATE_PROVINCE_CHOICES:
        for value, _label in options:
            if value in location_text:
                return value
    return ''


def extract_years_experience(screening_questions_text):
    match = YEARS_RE.search(screening_questions_text)
    return int(match.group(1)) if match else None


def build_internal_notes(row):
    lines = [
        f"Headline: {clean(row.get('Headline'))}",
        f"Applied to: {clean(row.get('Job Title'))} (Date Applied: {clean(row.get('Date Applied'))})",
        f"LinkedIn ATS stage at time of import: {clean(row.get('Current Stage'))} "
        "(this is LinkedIn's own status for that job posting -- unrelated to our pipeline stage)",
        f"Education: {clean(row.get('Education Degree'))} — {clean(row.get('Education Institution'))}",
        f"Screening questions (raw): {clean(row.get('Screening Questions'))}",
    ]
    return '\n'.join(lines)


def find_data_sheet(workbook):
    for sheet in workbook.worksheets:
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if first_row and clean(first_row[0]) == EXPECTED_HEADER:
            return sheet
    return None


@recruiter_required
def import_linkedin_candidates(request):
    if request.method != 'POST':
        return render(request, 'candidates/portal/import_candidates.html', {'active_nav': 'import'})

    upload = request.FILES.get('file')
    if not upload:
        return render(request, 'candidates/portal/import_candidates.html', {
            'active_nav': 'import', 'error': 'Choose a .xlsx file to upload.',
        })

    try:
        workbook = openpyxl.load_workbook(upload, data_only=True)
    except Exception:
        return render(request, 'candidates/portal/import_candidates.html', {
            'active_nav': 'import', 'error': 'Could not read that file -- is it a valid .xlsx export?',
        })

    sheet = find_data_sheet(workbook)
    if sheet is None:
        return render(request, 'candidates/portal/import_candidates.html', {
            'active_nav': 'import',
            'error': 'Could not find a worksheet with a "First Name" header row in this file.',
        })

    headers = [clean(c) for c in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    created, skipped_duplicate, errors = 0, 0, []

    for row_num, raw_row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = dict(zip(headers, raw_row))
        try:
            first_name = clean(row.get('First Name'))
            last_name = clean(row.get('Last Name'))
            linkedin_url = clean(row.get('Profile URL'))

            if not first_name or not last_name:
                errors.append((row_num, 'Missing First Name or Last Name.'))
                continue
            if not linkedin_url:
                errors.append((row_num, 'Missing Profile URL (LinkedIn URL is required).'))
                continue
            if Candidate.objects.filter(linkedin_url=linkedin_url).exists():
                skipped_duplicate += 1
                continue

            Candidate.objects.create(
                name=f'{first_name} {last_name}',
                email=clean(row.get('Email Address')),
                phone=clean(row.get('Phone Number')),
                linkedin_url=linkedin_url,
                current_company_name=clean(row.get('Current Company')),
                current_title=map_title(clean(row.get('Current Title'))),
                years_experience=extract_years_experience(clean(row.get('Screening Questions'))),
                state_province=find_state_province(clean(row.get('General Location'))),
                desired_ote='',
                open_to_relocation=None,
                internal_notes=build_internal_notes(row),
                resume='',
                source='linkedin',
            )
            created += 1
        except (IntegrityError, ValueError) as exc:
            errors.append((row_num, str(exc)))

    return render(request, 'candidates/portal/import_candidates.html', {
        'active_nav': 'import',
        'results': {'created': created, 'skipped_duplicate': skipped_duplicate, 'errors': errors},
    })
